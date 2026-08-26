
# excel_export.py -> Handles exporting attendance reports to Excel

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ==========================================================
# STYLES
# ==========================================================

REPORT_FONT = "Segoe UI"

TITLE_FONT = Font(
    name=REPORT_FONT,
    size=18,
    bold=True,
    color="1F4E78"
)

SUBTITLE_FONT = Font(
    name=REPORT_FONT,
    size=11,
    italic=True,
    color="595959"
)

HEADER_FONT = Font(
    name=REPORT_FONT,
    bold=True,
    color="FFFFFF",
    size=11
)

DATA_FONT = Font(
    name=REPORT_FONT,
    size=10
)


# ==========================================================
# FILLS
# ==========================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78"
)

ALT_FILL = PatternFill(
    fill_type="solid",
    start_color="F2F6FB"
)

WHITE_FILL = PatternFill(
    fill_type="solid",
    start_color="FFFFFF"
)


# ==========================================================
# BORDERS
# ==========================================================

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)


# ==========================================================
# ALIGNMENT
# ==========================================================

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)

LEFT = Alignment(
    horizontal="left",
    vertical="center"
)


# ==========================================================
# AUTO FIT COLUMNS
# ==========================================================

def autofit_columns(ws):

    for column in ws.columns:

        maximum = 0
        letter = get_column_letter(column[0].column)

        for cell in column:

            # Skip title rows in columns A-E
            if cell.row <= 4 and cell.column <= 5:
                continue

            try:

                if cell.value is not None:

                    val_len = len(str(cell.value))

                    if (
                        "yy" in str(cell.number_format)
                        or ":" in str(cell.number_format)
                    ):
                        val_len += 2

                    maximum = max(
                        maximum,
                        val_len
                    )

            except Exception:
                pass

        if maximum > 0:

            ws.column_dimensions[letter].width = (
                maximum + 4
            )

        else:

            ws.column_dimensions[letter].width = 15


# ==========================================================
# REPORT HEADER
# ==========================================================

def create_report_header(
    ws,
    title,
    statistics,
    alerts,
    table_cols_count
):

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    # Row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    # ------------------------------------------------------
    # TITLE
    # ------------------------------------------------------

    ws["A1"] = "Attendance Analysis System"

    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A1:E1")

    # ------------------------------------------------------
    # SUBTITLE
    # ------------------------------------------------------

    ws["A2"] = title

    ws["A2"].font = Font(
        name=REPORT_FONT,
        bold=True,
        size=14,
        color="333333"
    )

    ws.merge_cells("A2:E2")

    # ------------------------------------------------------
    # GENERATED DATE
    # ------------------------------------------------------

    ws["A3"] = (
        "Generated: "
        + datetime.now().strftime(
            "%d-%m-%Y %H:%M"
        )
    )

    ws["A3"].font = SUBTITLE_FONT

    ws.merge_cells("A3:E3")

    # ------------------------------------------------------
    # LOGO
    # ------------------------------------------------------

    logo_col = get_column_letter(
        max(table_cols_count + 5, 10)
    )

    try:

        from openpyxl.drawing.image import Image as ExcelImage

        logo_path = "static/logo.png"

        if os.path.exists(logo_path):

            img = ExcelImage(logo_path)

            img.width = 130
            img.height = 45

            ws.add_image(
                img,
                f"{logo_col}1"
            )

    except Exception:
        pass

    # ------------------------------------------------------
    # SUMMARY STATISTICS
    # ------------------------------------------------------

    stat_col_1 = get_column_letter(
        max(table_cols_count + 2, 7)
    )

    stat_col_2 = get_column_letter(
        max(table_cols_count + 3, 8)
    )

    stats_mapping = [
        (
            "Employees:",
            statistics.get("Employees", 0)
        ),
        (
            "Total Punches:",
            statistics.get("TotalPunches", 0)
        ),
        (
            "Alerts:",
            len(alerts)
        )
    ]

    for i, (label, value) in enumerate(
        stats_mapping,
        start=1
    ):

        cell_label = ws[
            f"{stat_col_1}{i}"
        ]

        cell_value = ws[
            f"{stat_col_2}{i}"
        ]

        cell_label.value = label

        cell_label.font = Font(
            name=REPORT_FONT,
            bold=True,
            color="1F4E78"
        )

        cell_label.alignment = Alignment(
            horizontal="right"
        )

        cell_value.value = value

        cell_value.font = DATA_FONT

        cell_value.alignment = Alignment(
            horizontal="left"
        )


# ==========================================================
# TABLE HEADER
# ==========================================================

def style_table_header(ws):

    ws.row_dimensions[5].height = 25

    for cell in ws[5]:

        cell.font = HEADER_FONT

        cell.fill = HEADER_FILL

        cell.alignment = CENTER

        cell.border = THIN_BORDER


# ==========================================================
# FINAL FORMATTING
# ==========================================================

def finish_sheet(ws):

    max_col_letter = get_column_letter(
        ws.max_column
    )

    max_row = ws.max_row

    for row in ws.iter_rows(
        min_row=6,
        max_col=ws.max_column,
        max_row=max_row
    ):

        ws.row_dimensions[
            row[0].row
        ].height = 20

        for cell in row:

            cell.font = DATA_FONT

            cell.border = THIN_BORDER

            if row[0].row % 2 == 0:

                cell.fill = ALT_FILL

            else:

                cell.fill = WHITE_FILL

    ws.freeze_panes = "A6"

    if max_row >= 5:

        ws.auto_filter.ref = (
            f"A5:{max_col_letter}{max_row}"
        )

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.oddFooter.center.text = (
        "Attendance Analysis System"
    )

    ws.oddFooter.right.text = (
        "Page &[Page] of &[Pages]"
    )

    autofit_columns(ws)


# ==========================================================
# ACTIVITY TRENDS CHART
# ==========================================================

def create_activity_chart(
    ws_chart,
    chart_labels,
    chart_data_in,
    chart_data_out
):

    # ------------------------------------------------------
    # HIDDEN CHART DATA
    # ------------------------------------------------------

    data_start_row = 30

    ws_chart.cell(
        row=data_start_row,
        column=1,
        value="Time"
    )

    ws_chart.cell(
        row=data_start_row,
        column=2,
        value="Entry (IN)"
    )

    ws_chart.cell(
        row=data_start_row,
        column=3,
        value="Exit (OUT)"
    )

    # ------------------------------------------------------
    # WRITE CHART DATA
    # ------------------------------------------------------

    for i in range(len(chart_labels)):

        row_number = data_start_row + 1 + i

        hour_text = chart_labels[i]

        # Convert "06:00" etc. into an actual Excel time.
        try:

            hour_value = datetime.strptime(
                hour_text,
                "%H:%M"
            ).time()

        except ValueError:

            # Fallback in case an unexpected value appears
            hour_value = hour_text

        time_cell = ws_chart.cell(
            row=row_number,
            column=1,
            value=hour_value
        )

        # Make Excel display the value as HH:MM
        time_cell.number_format = "hh:mm"

        ws_chart.cell(
            row=row_number,
            column=2,
            value=chart_data_in[i]
        )

        ws_chart.cell(
            row=row_number,
            column=3,
            value=chart_data_out[i]
        )

    # ------------------------------------------------------
    # CREATE BAR CHART
    # ------------------------------------------------------

    chart = BarChart()

    chart.type = "col"

    chart.style = 10

    chart.title = "Hourly Punch Activity"

    # Y axis
    chart.y_axis.title = "Number of Punches"

    # X axis
    chart.x_axis.title = "Hour of Day"

    # Put labels at the bottom
    chart.x_axis.tickLblPos = "low"

    # Show every category label
    chart.x_axis.tickLblSkip = 1

    # Chart dimensions
    chart.height = 13
    chart.width = 24

    # ------------------------------------------------------
    # DATA REFERENCE
    # ------------------------------------------------------

    data_ref = Reference(
        ws_chart,
        min_col=2,
        min_row=data_start_row,
        max_col=3,
        max_row=data_start_row + len(chart_labels)
    )

    # ------------------------------------------------------
    # CATEGORY REFERENCE
    # ------------------------------------------------------

    cats_ref = Reference(
        ws_chart,
        min_col=1,
        min_row=data_start_row + 1,
        max_row=data_start_row + len(chart_labels)
    )

    # Add IN / OUT data
    chart.add_data(
        data_ref,
        titles_from_data=True
    )

    # Add 06:00, 07:00, etc.
    chart.set_categories(cats_ref)

    chart.shape = 4

    # Add chart to worksheet
    ws_chart.add_chart(
        chart,
        "A5"
    )


# ==========================================================
# EXPORT FUNCTION
# ==========================================================

def export_excel(
    transactions,
    daily,
    overall,
    statistics,
    alerts,
    ml_data=None,
    chart_labels=None,
    chart_data_in=None,
    chart_data_out=None
):

    # ------------------------------------------------------
    # SAFE DEFAULTS
    # ------------------------------------------------------

    if ml_data is None:
        ml_data = []

    if chart_labels is None:
        chart_labels = []

    if chart_data_in is None:
        chart_data_in = []

    if chart_data_out is None:
        chart_data_out = []

    # ------------------------------------------------------
    # CREATE WORKBOOK
    # ------------------------------------------------------

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    # ======================================================
    # 1. ACTIVITY TRENDS
    # ======================================================

    if (
        chart_labels
        and chart_data_in
        and chart_data_out
    ):

        ws_chart = workbook.create_sheet(
            "Activity Trends"
        )

        ws_chart.sheet_view.showGridLines = False

        create_report_header(
            ws_chart,
            "Peak Entry & Exit Activity",
            statistics,
            alerts,
            table_cols_count=3
        )

        create_activity_chart(
            ws_chart,
            chart_labels,
            chart_data_in,
            chart_data_out
        )

    # ======================================================
    # 2. OVERALL SUMMARY
    # ======================================================

    ws = workbook.create_sheet(
        "Overall Summary"
    )

    for _ in range(4):
        ws.append([])

    ws.append([
        "Badge ID",
        "Working Days",
        "Total Punches",
        "First Entry",
        "Last Exit"
    ])

    create_report_header(
        ws,
        "Employee Summary",
        statistics,
        alerts,
        table_cols_count=5
    )

    style_table_header(ws)

    for row in overall:

        ws.append([
            row["BadgeID"],
            row["WorkingDays"],
            row["TotalPunches"],
            row["FirstEntry"],
            row["LastExit"]
        ])

    for column in (
        "A",
        "B",
        "C",
        "D",
        "E"
    ):

        for cell in ws[column]:

            if cell.row >= 6:

                cell.alignment = CENTER

                if column in ("D", "E"):

                    cell.number_format = (
                        "yyyy-mm-dd hh:mm:ss"
                    )

    finish_sheet(ws)

    # ======================================================
    # 3. DAILY SUMMARY
    # ======================================================

    ws = workbook.create_sheet(
        "Daily Summary"
    )

    for _ in range(4):
        ws.append([])

    ws.append([
        "Badge ID",
        "Date",
        "First IN",
        "Last OUT",
        "Punches"
    ])

    create_report_header(
        ws,
        "Daily Attendance Summary",
        statistics,
        alerts,
        table_cols_count=5
    )

    style_table_header(ws)

    for row in daily:

        ws.append([
            row["BadgeID"],
            row["Date"],
            row["FirstIn"],
            row["LastOut"],
            row["Punches"]
        ])

    for column in (
        "A",
        "B",
        "C",
        "D",
        "E"
    ):

        for cell in ws[column]:

            if cell.row >= 6:

                cell.alignment = CENTER

                if column == "B":

                    cell.number_format = (
                        "yyyy-mm-dd"
                    )

                if column in ("C", "D"):

                    cell.number_format = (
                        "hh:mm:ss"
                    )

    finish_sheet(ws)

    # ======================================================
    # 4. ALERTS
    # ======================================================

    ws = workbook.create_sheet(
        "Alerts"
    )

    for _ in range(4):
        ws.append([])

    ws.append([
        "Badge ID",
        "Date",
        "Date & Time",
        "Problem"
    ])

    create_report_header(
        ws,
        "Attendance Alerts",
        statistics,
        alerts,
        table_cols_count=4
    )

    style_table_header(ws)

    for row in alerts:

        ws.append([
            row["BadgeID"],
            row.get("Date", ""),
            row["Datetime"],
            row["Problem"]
        ])

    for column in (
        "A",
        "B",
        "C"
    ):

        for cell in ws[column]:

            if cell.row >= 6:

                cell.alignment = CENTER

                if column == "B":

                    cell.number_format = (
                        "yyyy-mm-dd"
                    )

                if column == "C":

                    cell.number_format = (
                        "yyyy-mm-dd hh:mm:ss"
                    )

    finish_sheet(ws)

    # ======================================================
    # 5. TRANSACTIONS
    # ======================================================

    ws = workbook.create_sheet(
        "Transactions"
    )

    for _ in range(4):
        ws.append([])

    ws.append([
        "Badge ID",
        "Date & Time",
        "Status",
        "Branch"
    ])

    create_report_header(
        ws,
        "Attendance Transactions",
        statistics,
        alerts,
        table_cols_count=4
    )

    style_table_header(ws)

    for row in transactions:

        ws.append([
            row["BadgeID"],
            row["Datetime"],
            row["Status"],
            row["Branch"]
        ])

    for cell in ws["B"]:

        if cell.row >= 6:

            cell.number_format = (
                "yyyy-mm-dd hh:mm:ss"
            )

            cell.alignment = CENTER

    for cell in ws["A"]:

        if cell.row >= 6:

            cell.alignment = CENTER

    for cell in ws["C"]:

        if cell.row >= 6:

            cell.alignment = CENTER

    finish_sheet(ws)

    # ======================================================
    # 6. STATISTICS
    # ======================================================

    ws = workbook.create_sheet(
        "Statistics"
    )

    for _ in range(4):
        ws.append([])

    ws.append([
        "Statistic",
        "Value"
    ])

    create_report_header(
        ws,
        "Attendance Statistics",
        statistics,
        alerts,
        table_cols_count=2
    )

    style_table_header(ws)

    for key, value in statistics.items():

        ws.append([
            key,
            value
        ])

    for cell in ws["B"]:

        if cell.row >= 6:

            cell.alignment = CENTER

    for cell in ws["A"]:

        if cell.row >= 6:

            cell.alignment = LEFT

    finish_sheet(ws)

    # ======================================================
    # 7. AI ANOMALIES
    # ======================================================

    if ml_data:

        ws = workbook.create_sheet(
            "Attendance Insights"
        )

        for _ in range(4):
            ws.append([])

        ws.append([
            "Badge ID",
            "Prediction",
            "Risk Level",
            "Anomaly Score",
            "Reasons"
        ])

        create_report_header(
            ws,
            "Attendance Insights",
            statistics,
            alerts,
            table_cols_count=5
        )

        style_table_header(ws)

        for row in ml_data:

            ws.append([
                row["BadgeID"],
                row["Prediction"],
                row["Risk"],
                row["Score"],
                row["Reasons"]
            ])

        for column in (
            "A",
            "B",
            "C",
            "D"
        ):

            for cell in ws[column]:

                if cell.row >= 6:

                    cell.alignment = CENTER

        for cell in ws["E"]:

            if cell.row >= 6:

                cell.alignment = LEFT

        finish_sheet(ws)

    # ======================================================
    # SAVE REPORT
    # ======================================================

    reports_folder = "reports"

    if not os.path.exists(
        reports_folder
    ):

        os.makedirs(
            reports_folder
        )

    filename = (
        "Attendance_Report_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".xlsx"
    )

    filepath = os.path.join(
        reports_folder,
        filename
    )

    workbook.save(filepath)

    return filepath
