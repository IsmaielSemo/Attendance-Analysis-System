# excel_export.py -> Handles the exporting to excel

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter

# ==========================================================
# STYLES (Professional Palette & Typography)
# ==========================================================
REPORT_FONT = "Segoe UI"  # Cleaner, modern font

TITLE_FONT = Font(name=REPORT_FONT, size=18, bold=True, color="1F4E78")  # Dark Blue
SUBTITLE_FONT = Font(name=REPORT_FONT, size=11, italic=True, color="595959")  # Slate Gray

HEADER_FONT = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name=REPORT_FONT, size=10)

# Fills
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78")  # Dark Blue
ALT_FILL = PatternFill(fill_type="solid", start_color="F2F6FB")  # Very Soft Blue/Gray
WHITE_FILL = PatternFill(fill_type="solid", start_color="FFFFFF")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# ==========================================================
# AUTO FIT COLUMNS
# ==========================================================
def autofit_columns(ws):
    for column in ws.columns:
        maximum = 0
        letter = get_column_letter(column[0].column)

        for cell in column:
            # Skip the long titles in columns A, B, and C so they don't stretch
            if cell.row <= 4 and cell.column <= 3:
                continue

            try:
                if cell.value is not None:
                    # Give a little extra padding if the cell holds a date/time
                    val_len = len(str(cell.value))
                    if "yy" in str(cell.number_format) or ":" in str(cell.number_format):
                        val_len += 2

                    maximum = max(maximum, val_len)
            except Exception:
                pass

        # Only apply a calculated width if we found data, otherwise set a safe default
        if maximum > 0:
            ws.column_dimensions[letter].width = maximum + 4
        else:
            ws.column_dimensions[letter].width = 15


# ==========================================================
# HEADER
# ==========================================================
def create_report_header(ws, title, statistics, alerts, table_cols_count):
    # Turn off Excel gridlines for a clean dashboard look
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Attendance Analysis System"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")  # Merge so it doesn't stretch column A

    # Subtitle
    ws["A2"] = title
    ws["A2"].font = Font(name=REPORT_FONT, bold=True, size=14, color="333333")
    ws.merge_cells("A2:C2")

    # Date Generated
    ws["A3"] = "Generated: " + datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["A3"].font = SUBTITLE_FONT
    ws.merge_cells("A3:C3")

    # ATTEMPT TO LOAD LOGO
    try:
        from openpyxl.drawing.image import Image as ExcelImage
        if os.path.exists("logo.png"):
            img = ExcelImage("logo.png")
            img.width = 120
            img.height = 40
            ws.add_image(img, "D1")
    except Exception:
        pass

    # Place the summary stats to the right side of the sheet dynamically
    stat_col_1 = get_column_letter(max(table_cols_count + 2, 5))
    stat_col_2 = get_column_letter(max(table_cols_count + 3, 6))

    stats_mapping = [
        ("Employees:", statistics.get("Employees", 0)),
        ("Total Punches:", statistics.get("TotalPunches", 0)),
        ("Alerts:", len(alerts))
    ]

    for i, (label, val) in enumerate(stats_mapping, start=1):
        cell_label = ws[f"{stat_col_1}{i}"]
        cell_val = ws[f"{stat_col_2}{i}"]

        cell_label.value = label
        cell_label.font = Font(name=REPORT_FONT, bold=True, color="1F4E78")
        cell_label.alignment = Alignment(horizontal="right")

        cell_val.value = val
        cell_val.font = DATA_FONT
        cell_val.alignment = Alignment(horizontal="left")


# ==========================================================
# TABLE HEADER
# ==========================================================
def style_table_header(ws):
    ws.row_dimensions[5].height = 25  # Make header row taller

    for cell in ws[5]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


# ==========================================================
# FINAL FORMATTING
# ==========================================================
def finish_sheet(ws):
    max_col_letter = get_column_letter(ws.max_column)
    max_row = ws.max_row

    for row in ws.iter_rows(min_row=6, max_col=ws.max_column, max_row=max_row):
        ws.row_dimensions[row[0].row].height = 20

        for cell in row:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if row[0].row % 2 == 0:
                cell.fill = ALT_FILL
            else:
                cell.fill = WHITE_FILL

    ws.freeze_panes = "A6"

    if max_row >= 5:
        ws.auto_filter.ref = f"A5:{max_col_letter}{max_row}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.oddFooter.center.text = "Attendance Analysis System"
    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"

    autofit_columns(ws)


# ==========================================================
# EXPORT FUNCTION
# ==========================================================
def export_excel(transactions, daily, overall, statistics, alerts, ml_data=[]):
    workbook = Workbook()
    workbook.remove(workbook.active)

    # 1. OVERALL SUMMARY
    ws = workbook.create_sheet("Overall Summary")
    ws.append([]);
    ws.append([]);
    ws.append([]);
    ws.append([])
    ws.append(["Badge ID", "Working Days", "Total Punches", "First Entry", "Last Exit"])
    create_report_header(ws, "Employee Summary", statistics, alerts, table_cols_count=5)
    style_table_header(ws)
    for row in overall:
        ws.append([row["BadgeID"], row["WorkingDays"], row["TotalPunches"], row["FirstEntry"], row["LastExit"]])
    for column in ("A", "B", "C", "D", "E"):
        for cell in ws[column]:
            if cell.row >= 6:
                cell.alignment = CENTER
                if column in ("D", "E"):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
    finish_sheet(ws)

    # 2. DAILY SUMMARY
    ws = workbook.create_sheet("Daily Summary")
    ws.append([]);
    ws.append([]);
    ws.append([]);
    ws.append([])
    ws.append(["Badge ID", "Date", "First IN", "Last OUT", "Punches"])
    create_report_header(ws, "Daily Attendance Summary", statistics, alerts, table_cols_count=5)
    style_table_header(ws)
    for row in daily:
        ws.append([row["BadgeID"], row["Date"], row["FirstIn"], row["LastOut"], row["Punches"]])
    for column in ("A", "B", "C", "D", "E"):
        for cell in ws[column]:
            if cell.row >= 6:
                cell.alignment = CENTER
                if column in ("C", "D"):
                    cell.number_format = "hh:mm:ss"
    finish_sheet(ws)

    # 3. ALERTS
    ws = workbook.create_sheet("Alerts")
    ws.append([]);
    ws.append([]);
    ws.append([]);
    ws.append([])
    ws.append(["Badge ID", "Date", "Date & Time", "Problem"])
    create_report_header(ws, "Attendance Alerts", statistics, alerts, table_cols_count=4)
    style_table_header(ws)
    for row in alerts:
        ws.append([row["BadgeID"], row.get("Date", ""), row["Datetime"], row["Problem"]])
    for column in ("A", "B", "C"):
        for cell in ws[column]:
            if cell.row >= 6:
                cell.alignment = CENTER
                if column == "C":
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
    finish_sheet(ws)

    # 4. TRANSACTIONS
    ws = workbook.create_sheet("Transactions")
    ws.append([]);
    ws.append([]);
    ws.append([]);
    ws.append([])
    ws.append(["Badge ID", "Date & Time", "Status", "Branch"])
    create_report_header(ws, "Attendance Transactions", statistics, alerts, table_cols_count=4)
    style_table_header(ws)
    for row in transactions:
        ws.append([row["BadgeID"], row["Datetime"], row["Status"], row["Branch"]])
    for cell in ws["B"]:
        if cell.row >= 6:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
            cell.alignment = CENTER
    for cell in ws["A"]:
        if cell.row >= 6: cell.alignment = CENTER
    for cell in ws["C"]:
        if cell.row >= 6: cell.alignment = CENTER
    finish_sheet(ws)

    # 5. STATISTICS
    ws = workbook.create_sheet("Statistics")
    ws.append([]);
    ws.append([]);
    ws.append([]);
    ws.append([])
    ws.append(["Statistic", "Value"])
    create_report_header(ws, "Attendance Statistics", statistics, alerts, table_cols_count=2)
    style_table_header(ws)
    for key, value in statistics.items():
        ws.append([key, value])
    for cell in ws["B"]:
        if cell.row >= 6: cell.alignment = CENTER
    for cell in ws["A"]:
        if cell.row >= 6: cell.alignment = LEFT
    finish_sheet(ws)

    # 6. AI ANOMALIES (Now placed last)
    if ml_data:
        ws = workbook.create_sheet("AI Anomalies")
        ws.append([]);
        ws.append([]);
        ws.append([]);
        ws.append([])
        ws.append(["Badge ID", "Prediction", "Risk Level", "Anomaly Score", "Reasons"])
        create_report_header(ws, "AI Anomaly Detection", statistics, alerts, table_cols_count=5)
        style_table_header(ws)
        for row in ml_data:
            ws.append([row["BadgeID"], row["Prediction"], row["Risk"], row["Score"], row["Reasons"]])
        for column in ("A", "B", "C", "D"):
            for cell in ws[column]:
                if cell.row >= 6: cell.alignment = CENTER
        for cell in ws["E"]:
            if cell.row >= 6: cell.alignment = LEFT
        finish_sheet(ws)

    reports_folder = "reports"
    if not os.path.exists(reports_folder):
        os.makedirs(reports_folder)

    filename = "Attendance_Report_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
    filepath = os.path.join(reports_folder, filename)

    workbook.save(filepath)
    return filepath
